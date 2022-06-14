# coding=utf8
from selenium import webdriver
from selenium_stealth import stealth
from selenium.webdriver.common.keys import Keys
from selenium.webdriver import ActionChains
from openpyxl import Workbook
import time

driver = webdriver.Chrome()
driver.get("https://domains.google/")
driver.maximize_window()

stealth(driver,
        languages=["en-US", "en"],
        vendor="Google Inc.",
        platform="Win32",
        webgl_vendor="Intel Inc.",
        renderer="Intel Iris OpenGL Engine",
        fix_hairline=True,
        )

# jinleedvp0@gmail.com
# January01)!
# Get a new domain
# driver.find_element_by_xpath('').click()

driver.find_element_by_xpath('/html/body/div[2]/section[1]/div/section/div[1]/div[3]/form/input').click()
driver.find_element_by_xpath('/html/body/div[2]/section[1]/div/section/div[1]/div[3]/form/input').send_keys("test")
driver.find_element_by_xpath('/html/body/div[2]/section[1]/div/section/div[1]/div[3]/form/button').click()
time.sleep(5)

alphabet = ["a", "b", "c", "d", "e", "f", "g", "h", "i", "j", "k", "l", "m", "n", "o", "p", "q", "r", "s", "t", "u", "v", "w", "x", "y", "z"]
consonants = ["b", "c", "d", "f", "g", "h", "j", "k", "l", "m", "n", "p", "q", "r", "s", "t", "v", "w", "x", "y", "z"]
vowels = ["a", "e", "i", "o", "u"]
twosyl = [
"pecke", "pecks", "pecky", "pedal", "pedes", "pedos", "pedro", "peece", "peeks", "peels", "peens", "peeoy", "peepe", "peeps", "peers", "peery", "peeve", "peggy", "peghs", "pegos", "peine", "peins", "peise", "peize", "pekan", "pekes", "pekin", "pekoe", "pelas", "pelau", "peles", "pelfs", "pells", "pelma", "pelon", "pelta", "pelts", "penal", "pence", "pends", "pendu", "pened", "penes", "pengo", "penie", "penis", "penks", "penna", "penne", "penni", "penny", "pense", "pents", "peons", "peony", "pepla", "pepos", "peppy", "pepsi", "perai", "perce", "perch", "percs", "perdu", "perdy", "perea", "peres", "peril", "peris", "perks", "perky", "perle", "perms", "perns", "perps", "perry", "perse", "perst", "perts", "perve", "pervo", "pervs", "pervy", "pesky", "pesos", "pesto", "pests", "pesty", "petal", "petar", "peter", "petit", "petre", "petti", "petto", "petty", "pewee", "pewit", "peyse", "phage", "phang", "phare", "phase", "pheer", "phene", "pheon", "phese", "phial", "phish", "phlox", "phoca", "phohs", "phone", "phono", "phons", "phony", "photo", "phots", "phpht", "phuts", "phyla", "phyle", "piani", "piano", "pians", "pibal", "pical", "picas", "piccy", "pichi", "picks", "picky", "picot", "picra", "picul", "piece", "pieds", "piend", "piers", "piert", "pieta", "piets", "piety", "piezo", "piggy", "pight", "pigmy", "piing", "pikas", "pikau", "piked", "piker", "pikes", "pikey", "pikis", "pikul", "pilaf", "pilao", "pilar", "pilau", "pilaw", "pilch", "pilea", "piled", "pilei", "piler", "piles", "pilis", "pills", "pilot", "pilow", "pilum", "pilus", "pimas", "pimps", "pinas", "pinch", "pined", "piner", "pines", "piney", "pingo", "pings", "pinko", "pinks", "pinky", "pinna", "pinny", "pinon", "pinot", "pinta", "pinto", "pints", "pinup", "pions", "piony", "pious", "pioye", "pioys", "pipal", "pipas", "piped", "piper", "pipes", "pipet", "pipis", "pipit", "pippy", "pipul", "pique", "pirai", "pirks", "pirls", "pirns", "pirog", "pirri", "pisco", "pises", "pisky", "pisos", "pissy", "piste", "pitas", "pitch", "piths", "pithy", "piton", "pitot", "pitta", "piums", "pivot", "pixel", "pixes", "pixie", "pized", "pizes", "pizza", "plaas", "place", "plack", "plage", "plaid", "plain", "plait", "plane", "plank", "plans", "plant", "plaps", "plash", "plasm", "plast", "plate", "plats", "platy", "playa", "plays", "plaza", "plead", "pleas", "pleat", "plebe", "plebs", "plein", "plena", "pleno", "pleon", "plesh", "plews", "plica", "plied", "plier", "plies", "plims", "pling", "plink", "ploat", "plock", "plods", "plong", "plonk", "plook", "plops", "plots", "plotz", "plouk", "plows", "ploye", "ploys", "pluck", "plues", "pluff", "plugs", "pluke", "plumb", "plume", "plump", "plums", "plumy", "plunk", "pluot", "plush", "plyer", "poach", "poaka", "poake", "poboy", "pocks", "pocky", "podal", "poddy", "podex", "podge", "podgy", "podia", "poems", "poeps", "poesy", "poete", "poets", "pogey", "pogge", "pogos", "poilu", "poind", "point", "poise", "pokal", "poked", "poker", "pokes", "pokey", "pokie", "polar", "poled", "poler", "poles", "poley", "polio", "polis", "polje", "polka", "polks", "pollo", "polls", "polly", "polos", "polts", "polyp", "polys", "pomas", "pombe", "pomes", "pommy", "pomos", "pomps", "ponce", "poncy", "ponds", "pones", "poney", "ponga", "pongo", "pongs", "pongy", "ponks", "ponts", "ponty", "ponzu", "pooch", "poods", "pooed", "poofs", "poofy", "poohs", "pooja", "pooka", "pooks", "pools", "poons", "poops", "poopy", "poori", "poort", "poots", "poove", "poovy", "popes", "popos", "poppa", "poppy", "popsy", "popup", "porae", "poral", "porch", "pored", "porer", "pores", "porge", "porgy", "porin", "porks", "porky", "porno", "porns", "porny", "porta", "porte", "ports", "porty", "posed", "poser", "poses", "posey", "posho", "posit", "posse", "poste", "posts", "potae", "potch", "poted", "potes", "potin", "potoo", "potsy", "potto", "potts", "potty", "potus", "pouch", "pouff", "poufs", "poufy", "pouis", "pouke", "pouks", "poule", "poulp", "poult", "pound", "poupe", "poupt", "pours", "pouts", "pouty", "powan", "power", "powin", "pownd", "powns", "powny", "powre", "poxed", "poxes", "poynt", "poyou", "poyse", "pozzy", "praam", "prads", "prahu", "prams", "prana", "prang", "prank", "praos", "prase", "prate", "prats", "pratt", "praty", "praus", "prawn", "prays", "predy", "preed", "preen", "prees", "preif", "prems", "premy", "prent", "preop", "preps", "presa", "prese", "press", "prest", "preve", "prexy", "preys", "prial", "price", "prick", "pricy", "pride", "pried", "prief", "prier", "pries", "prigs", "prill", "prima", "prime", "primi", "primo", "primp", "prims", "primy", "prink", "print", "prion", "prior", "prise", "prism", "priss", "privy", "prize", "proas", "probe", "probs", "proby", "prods", "proem", "profs", "progs", "proin", "proke", "prole", "proll", "promo", "proms", "prone", "prong", "pronk", "proof", "props", "prore", "prose", "proso", "pross", "prost", "prosy", "proud", "proul", "prove", "prowl", "prows", "proxy", "proyn", "prude", "prune", "prunt", "pruta", "pryer", "pryse", "psalm", "pseud", "pshaw", "psion", "psoae", "psoai", "psoas", "psora", "psych", "psyop", "ptype", "pubes", "pubic", "pubis", "pucan", "pucer", "puces", "pucka", "pucks", "puddy", "pudge", "pudgy", "pudic", "pudor", "pudsy", "pudus", "puers", "puffs", "puffy", "puggy", "pugil", "puhas", "pujah", "pujah", "pujas", "pukas", "puked", "puker", "pukes", "pukey", "pukka", "pukus", "pulao", "pulas", "puled", "puler", "pules", "pulik", "pulis", "pulka", "pulks", "pulli", "pulls", "pulmo", "pulps", "pulpy", "pulse", "pulus", "pumas", "pumie", "pumps", "pumpy", "punas", "punce", "punch", "punga", "pungs", "punim", "punji", "punka", "punks", "punky", "punny", "punto", "punts", "punty", "pupae", "pupal", "pupas", "pupil", "puppo", "puppy", "pupus", "purda", "pured", "puree", "purer", "pures", "purge", "purim", "purin", "puris", "purls", "puros", "purpy", "purrs", "purse", "pursy", "purty", "puses", "pushy", "pusle", "pussy", "puter", "putid", "putin", "puton", "putti", "putto", "putts", "putty", "puzel", "pyats", "pyets", "pygal", "pygmy", "pyins", "pylon", "pyned", "pynes", "pyoid", "pyots", "pyral", "pyran", "pyres", "pyrex", "pyric", "pyros", "pyxed", "pyxes", "pyxie", "pyxis", "pzazz"

]
test = ["phese", "test"]

wb = Workbook() # new workbook
ws = wb.active # call a sheet
ws_y=1

# for char1 in consonants:
#     for char2 in vowels:
#         for char3 in consonants:
#             for char4 in vowels:
#                 for char5 in consonants:
#                     nsearch = char1 + char2 + char3 + char4 + char5 + ".com"
for char1 in twosyl:
    nsearch = char1 + ".com"
    driver.find_element_by_xpath('//*[@id="mat-input-0"]').click()
    ActionChains(driver).send_keys(Keys.BACKSPACE).send_keys(Keys.BACKSPACE).send_keys(Keys.BACKSPACE).send_keys(Keys.BACKSPACE).send_keys(Keys.BACKSPACE).send_keys(Keys.BACKSPACE).send_keys(Keys.BACKSPACE).send_keys(Keys.BACKSPACE).send_keys(Keys.BACKSPACE).send_keys(Keys.BACKSPACE).send_keys(Keys.BACKSPACE).send_keys(Keys.BACKSPACE).send_keys(Keys.BACKSPACE).send_keys(Keys.BACKSPACE).send_keys(Keys.BACKSPACE).perform()# delete previous search text
    driver.find_element_by_xpath('//*[@id="mat-input-0"]').send_keys(nsearch)# combination search text
    driver.find_element_by_xpath('//*[@id="domainsAppContent"]/main/dreg-router-outlet/div/search-root/responsive-centered-container/div/div/search-bar-container/findy-bar-container/div/search-bar/findy-bar/form/button[1]').click()# search
    time.sleep(5)
    try:
        adomain = driver.find_element_by_xpath('//*[@id="mat-tab-content-0-0"]/div/related-search-results/exact-match-card/div/search-result-card-header/search-result-card-header-desktop/mat-card/div/div[2]/div/search-result-card-domain-name/span').text
        ws.cell(column=1, row=ws_y, value=adomain)
        aprice = driver.find_element_by_xpath('//*[@id="mat-tab-content-0-0"]/div/related-search-results/exact-match-card/div/search-result-card-header/search-result-card-header-desktop/mat-card/div/div[3]/row-price/domain-price/span[1]').text
        ws.cell(column=2, row=ws_y, value=aprice)
        wb.save("1.xlsx")
        ws_y += 1
        print(adomain + "saved")
    except:
        pass
                    # try:# Add to favorites
                    #     driver.find_element_by_xpath('//*[@id="mat-tab-content-0-0"]/div/related-search-results/exact-match-card/div/search-result-card-header/search-result-card-header-desktop/mat-card/div/div[4]/row-add-to-favorites/div/button').click()
                    # except:
                    #     pass
