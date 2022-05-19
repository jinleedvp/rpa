from openpyxl import load_workbook

wb = load_workbook("1.xlsx")
ws4 = wb["range"]

for y in range(1,ws4.max_column+1):
    for x in range(1,ws4.max_row+1):
        print(ws4.cell(column=x, row=y).value, end="  ")
    print()
