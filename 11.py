from openpyxl import Workbook
from random import *

wb = Workbook() # 새 워크북 생성
ws = wb.active # 현재 활성화된 sheet 가져옴
ws.title = "jinlee" # sheet 이름 변경
ws.sheet_properties.tabColor = "ff66ff"
ws2 = wb.create_sheet("jinlee2", 2)
ws2 = wb["jinlee2"]
ws2["A1"] = "Test"
ws2.cell(column=1, row=2, value=1)
ws3 = wb.copy_worksheet(ws2)

ws4 = wb.create_sheet("range")
index = 1
for y in range(1,11):
    for x in range(1,11):
        ws4.cell(column=x, row=y, value=index)
        index += 1


ws5 = wb.create_sheet("apx")
ws5.append(["국어", "영어", "수학"])
for i in range(1,11):
    ws.append([i,randint(1,100),randint(1,100)])


print(wb.sheetnames)
print(ws2["A1"].value)
print(ws2.cell(column=1, row=1).value)

wb.save("1.xlsx")
wb.close()
