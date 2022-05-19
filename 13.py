from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC

from openpyxl import Workbook

import time

driver = webdriver.Chrome()
driver.get("https://search.dca.ca.gov/")
driver.maximize_window()

# driver.find_element_by_link_text("문자")
# driver.find_elements_by_link_text("문자")[0]
driver.find_element_by_xpath('//*[@id="licenseType"]').send_keys("Licensed Acupuncturist")
driver.find_element_by_xpath('//*[@id="licenseNumber"]').send_keys("1")
driver.find_element_by_xpath('//*[@id="srchSubmitHome"]').click()
try:
    WebDriverWait(driver, 10).until(EC.presence_of_element_located((By.XPATH, '//*[@id="0"]/footer/ul[1]/li[1]/h3')))
finally:
    None

wb = Workbook() # new workbook
ws = wb.active # call a sheet
ws_y=9626

for i in range(9626,19200):#19195
    driver.find_element_by_xpath('//*[@id="licenseNumber"]').clear()
    driver.find_element_by_xpath('//*[@id="licenseNumber"]').send_keys(i)
    driver.find_element_by_xpath('//*[@id="srchSubmit"]').click()
    licnum = i
    ws.cell(column=1, row=ws_y, value=licnum)
    try:
        WebDriverWait(driver, 10).until(EC.presence_of_element_located((By.XPATH, '//*[@id="0"]/footer/ul[1]/li[1]/h3')))# Name
        driver.find_element_by_xpath('//*[@id="mD0"]').click()# More Detail
        time.sleep(1)
        driver.switch_to.window(driver.window_handles[1])# tab switch
        WebDriverWait(driver, 10).until(EC.presence_of_element_located((By.XPATH, '//*[@id="clntType"]')))


        try:
            name = driver.find_element_by_xpath('//*[@id="name"]').text
            ws.cell(column=2, row=ws_y, value=name)
        except:
            pass

        try:
            lictyp = driver.find_element_by_xpath('//*[@id="licType"]').text
            ws.cell(column=3, row=ws_y, value=lictyp)
        except:
            pass

        try:
            licstu = driver.find_element_by_xpath('//*[@id="primaryStatus"]').text
            ws.cell(column=4, row=ws_y, value=licstu)
        except:
            pass

        try:
            licstu2 = driver.find_element_by_xpath('//*[@id="secondaryStatus"]').text
            ws.cell(column=5, row=ws_y, value=licstu2)
        except:
            pass

        try:
            prename = driver.find_element_by_xpath('//*[@id="prevName"]').text
            ws.cell(column=6, row=ws_y, value=prename)
        except:
            pass

        try:
            address = driver.find_element_by_xpath('//*[@id="address"]/p[2]').text
            ws.cell(column=7, row=ws_y, value=address)
        except:
            pass

        try:
            liciss = driver.find_element_by_xpath('//*[@id="issueDate"]').text
            ws.cell(column=8, row=ws_y, value=liciss)
        except:
            pass

        try:
            licexp = driver.find_element_by_xpath('//*[@id="expDate"]').text
            ws.cell(column=9, row=ws_y, value=licexp)
        except:
            pass

        wb.save("1.xlsx")
        time.sleep(1)
        driver.close()
        time.sleep(1)
        driver.switch_to.window(driver.window_handles[0])
        time.sleep(1)
        ws_y += 1
        # input("Press Enter to continue...")
    except:
        wb.save("1.xlsx")
        time.sleep(1)
        driver.switch_to.window(driver.window_handles[0])
        time.sleep(1)
        ws_y += 1
        # input("Press Enter to continue...")
